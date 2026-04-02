"""XLSX Engine — Spreadsheet generation via openpyxl."""

from pathlib import Path

from generate import register_engine


@register_engine('xlsx')
class XlsxEngine:
    def generate(self, data, template_path, output_path, schema=None):
        """
        Generate XLSX from data.

        Expected data formats:
        1. {"sheets": [{"name": "Sheet1", "headers": [...], "rows": [...], "formulas": {...}}]}
        2. {"items": [...]}  (single sheet with auto-headers)
        3. {"headers": [...], "rows": [...]}  (single sheet explicit)
        """
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter
        except ImportError:
            raise RuntimeError("openpyxl required: pip install openpyxl")

        Path(output_path).parent.mkdir(parents=True, exist_ok=True)
        wb = Workbook()

        # Brand colors
        header_fill = PatternFill(start_color='16213E', end_color='16213E', fill_type='solid')
        header_font = Font(name='Calibri', bold=True, color='FFFFFF', size=11)
        data_font = Font(name='Calibri', size=11)
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )

        def style_header(ws, row=1, col_count=1):
            for col in range(1, col_count + 1):
                cell = ws.cell(row=row, column=col)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center')
                cell.border = thin_border

        def style_data(ws, start_row, end_row, col_count):
            for row in range(start_row, end_row + 1):
                for col in range(1, col_count + 1):
                    cell = ws.cell(row=row, column=col)
                    cell.font = data_font
                    cell.border = thin_border

        def auto_width(ws, col_count):
            for col in range(1, col_count + 1):
                max_len = 0
                letter = get_column_letter(col)
                for row in ws.iter_rows(min_col=col, max_col=col):
                    for cell in row:
                        if cell.value:
                            max_len = max(max_len, len(str(cell.value)))
                ws.column_dimensions[letter].width = min(max_len + 4, 40)

        # Multi-sheet format
        if 'sheets' in data:
            wb.remove(wb.active)
            for sheet_def in data['sheets']:
                ws = wb.create_sheet(title=sheet_def.get('name', 'Sheet'))
                headers = sheet_def.get('headers', [])
                rows = sheet_def.get('rows', [])
                formulas = sheet_def.get('formulas', {})

                # Write headers
                if headers:
                    for col_idx, h in enumerate(headers, 1):
                        ws.cell(row=1, column=col_idx, value=h)
                    style_header(ws, 1, len(headers))

                # Write data rows
                for row_idx, row_data in enumerate(rows, 2):
                    for col_idx, val in enumerate(row_data, 1):
                        ws.cell(row=row_idx, column=col_idx, value=val)
                    style_data(ws, row_idx, row_idx, len(headers) or len(row_data))

                # Write formulas
                for cell_ref, formula in formulas.items():
                    ws[cell_ref] = formula

                if headers:
                    auto_width(ws, len(headers))

        # Single sheet from items
        elif 'items' in data or 'records' in data:
            ws = wb.active
            ws.title = data.get('title', 'Data')
            items = data.get('items', data.get('records', []))

            if not items:
                wb.save(str(output_path))
                return

            headers = list(items[0].keys())

            for col_idx, h in enumerate(headers, 1):
                ws.cell(row=1, column=col_idx, value=h)
            style_header(ws, 1, len(headers))

            for row_idx, item in enumerate(items, 2):
                for col_idx, h in enumerate(headers, 1):
                    ws.cell(row=row_idx, column=col_idx, value=item.get(h, ''))
                style_data(ws, row_idx, row_idx, len(headers))

            auto_width(ws, len(headers))

        # Single sheet from headers + rows
        elif 'headers' in data and 'rows' in data:
            ws = wb.active
            ws.title = data.get('title', 'Data')
            headers = data['headers']

            for col_idx, h in enumerate(headers, 1):
                ws.cell(row=1, column=col_idx, value=h)
            style_header(ws, 1, len(headers))

            for row_idx, row_data in enumerate(data['rows'], 2):
                for col_idx, val in enumerate(row_data, 1):
                    ws.cell(row=row_idx, column=col_idx, value=val)
                style_data(ws, row_idx, row_idx, len(headers))

            auto_width(ws, len(headers))

        else:
            raise ValueError("XLSX data must have 'sheets', 'items'/'records', or 'headers'+'rows'")

        wb.save(str(output_path))
